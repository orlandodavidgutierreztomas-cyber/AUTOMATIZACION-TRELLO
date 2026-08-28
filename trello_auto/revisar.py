#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
============================================================================
 REVISAR — el cuadro de verificacion del mapeo, en Excel.
============================================================================

EL PROBLEMA
-----------
Despues de sincronizar hay que revisar a que familia y a que lista va cada
actividad. Hacerlo escribiendo a mano dentro de un JSON es pedir una errata:
un acento de mas, una tilde de menos, y esa actividad deja de encontrar su
lista sin que nadie se entere.

LA SOLUCION
-----------
Un Excel donde NO SE ESCRIBE: se ELIGE. Cada celda editable es una lista
desplegable con las opciones validas, sacadas de tu propia configuracion y
de los nombres reales de las listas de tu tablero. No se puede escribir algo
que no exista.

EL CICLO
--------
  1. "Sincronizar"  genera  mapeo/revisar_mapeo.xlsx
  2. Lo descargas, revisas y eliges en los desplegables. La columna REVISAR
     te marca las filas que piden atencion, para no mirar las 75.
  3. Lo vuelves a subir al repositorio, a la misma ruta.
  4. "Aplicar mapeo"  lo lee, valida y reescribe mapeo.json.

Si algo no cuadra, el paso 4 falla con un mensaje claro y NO escribe nada:
mejor no aplicar que aplicar a medias.

USO
---
    python -m trello_auto.revisar --generar
    python -m trello_auto.revisar --aplicar
    python -m trello_auto.revisar --aplicar --dry-run
============================================================================
"""

from __future__ import annotations

import argparse
import json
import os
import sys

from . import ajustes

RUTA_EXCEL = ajustes.RAIZ / "mapeo" / "revisar_mapeo.xlsx"
HOJA_MAPEO = "MAPEO"
HOJA_OPCIONES = "_opciones"

# Columna -> (encabezado, ancho, editable)
COLUMNAS = [
    ("CLAVE (no editar)", 34, False),
    ("ACTIVIDAD", 56, False),
    ("FAMILIA", 16, True),
    ("LISTA DESTINO", 42, True),
    ("TIENE PLANTILLA", 16, False),
    ("VECES EN EL PLAN", 17, False),
    ("REVISAR", 26, False),
]


def _opciones_familia() -> list:
    return [f for f in ajustes.FAMILIAS]


def _opciones_lista(mapeo: dict) -> list:
    """Nombres de lista entre los que elegir: los reales del tablero primero."""
    opciones = list(mapeo.get("listas_del_tablero") or [])
    for familia in ajustes.FAMILIAS:
        lista = ajustes.lista_de_familia(familia)
        if lista and lista not in opciones:
            opciones.append(lista)
    return opciones


def _motivo_revision(fila: dict) -> str:
    """Por que esta fila merece una mirada. Vacio = no hace falta tocarla."""
    motivos = []
    if fila.get("familia") == ajustes.familia_por_defecto():
        motivos.append("cayo en el descarte")
    if not fila.get("tiene_plantilla"):
        motivos.append("sin plantilla")
    return " · ".join(motivos)


def generar() -> str:
    """Escribe el cuadro de verificacion a partir de mapeo.json."""
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation

    mapeo = json.loads(ajustes.ARCHIVO_MAPEO.read_text(encoding="utf-8"))
    actividades = mapeo.get("actividades") or {}
    if not actividades:
        raise SystemExit(
            "ERROR: mapeo.json no tiene actividades.\n"
            "Corre primero 'Sincronizar cronograma y tablero'."
        )

    familias = _opciones_familia()
    listas = _opciones_lista(mapeo)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = HOJA_MAPEO

    # Las opciones viven en una hoja aparte: asi los desplegables no tienen
    # el limite de 255 caracteres y aguantan nombres largos con emojis.
    op = wb.create_sheet(HOJA_OPCIONES)
    for i, f in enumerate(familias, 1):
        op.cell(i, 1).value = f
    for i, lst in enumerate(listas, 1):
        op.cell(i, 2).value = lst
    op.sheet_state = "hidden"

    # Encabezado
    azul = PatternFill("solid", fgColor="1F4E79")
    for c, (titulo, ancho, _) in enumerate(COLUMNAS, 1):
        cel = ws.cell(1, c)
        cel.value = titulo
        cel.font = Font(bold=True, color="FFFFFF", size=10)
        cel.fill = azul
        cel.alignment = Alignment(vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(c)].width = ancho
    ws.freeze_panes = "C2"
    ws.row_dimensions[1].height = 30

    gris = Font(color="808080")
    ambar = PatternFill("solid", fgColor="FFF3CD")

    filas = sorted(actividades.items(),
                   key=lambda kv: (kv[1].get("familia", ""), kv[1].get("actividad", "")))
    for r, (clave, datos) in enumerate(filas, 2):
        motivo = _motivo_revision(datos)
        valores = [
            clave,
            datos.get("actividad", ""),
            datos.get("familia", ""),
            datos.get("lista", ""),
            "si" if datos.get("tiene_plantilla") else "NO",
            datos.get("veces_en_el_plan", 0),
            motivo,
        ]
        for c, valor in enumerate(valores, 1):
            cel = ws.cell(r, c)
            cel.value = valor
            if not COLUMNAS[c - 1][2]:          # columnas de solo lectura
                cel.font = gris
        if motivo:
            for c in range(1, len(COLUMNAS) + 1):
                ws.cell(r, c).fill = ambar

    ultima = len(filas) + 1

    # Los desplegables: se elige, no se escribe
    dv_fam = DataValidation(
        type="list", allow_blank=False,
        formula1=f"={HOJA_OPCIONES}!$A$1:$A${len(familias)}",
        showDropDown=False, errorTitle="Familia no valida",
        error="Elige una de la lista. No escribas a mano.")
    dv_lst = DataValidation(
        type="list", allow_blank=False,
        formula1=f"={HOJA_OPCIONES}!$B$1:$B${len(listas)}",
        showDropDown=False, errorTitle="Lista no valida",
        error="Elige una de las listas de tu tablero.")
    ws.add_data_validation(dv_fam)
    ws.add_data_validation(dv_lst)
    dv_fam.add(f"C2:C{ultima}")
    dv_lst.add(f"D2:D{ultima}")
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNAS))}{ultima}"

    # Instrucciones, debajo de la tabla
    ayuda = [
        "",
        "COMO SE USA",
        "1. Revisa las filas marcadas en ambar: son las que piden atencion.",
        "2. En FAMILIA y LISTA DESTINO elige del desplegable. No escribas a mano.",
        "3. No toques CLAVE ni ACTIVIDAD: son las que enlazan con el cronograma.",
        "4. Guarda, sube este archivo al repositorio en la misma ruta,",
        "   y corre el workflow 'Aplicar mapeo'.",
        "",
        "Si no hay nada que corregir, no hace falta hacer nada.",
    ]
    for i, texto in enumerate(ayuda):
        cel = ws.cell(ultima + 2 + i, 2)
        cel.value = texto
        cel.font = Font(bold=(i == 1), size=10)

    os.makedirs(os.path.dirname(RUTA_EXCEL), exist_ok=True)
    wb.save(RUTA_EXCEL)

    por_revisar = sum(1 for _, d in filas if _motivo_revision(d))
    print(f"Cuadro de verificacion: {RUTA_EXCEL}")
    print(f"  {len(filas)} actividades · {por_revisar} marcadas para revisar")
    print(f"  Opciones de familia: {', '.join(familias)}")
    print(f"  Opciones de lista:   {len(listas)}")
    return str(RUTA_EXCEL)


def aplicar(dry_run: bool = False) -> int:
    """Lee el cuadro revisado y reescribe mapeo.json. Valida antes de escribir."""
    import openpyxl

    if not os.path.exists(RUTA_EXCEL):
        raise SystemExit(
            f"ERROR: no encuentro {RUTA_EXCEL}.\n"
            f"Corre 'Sincronizar' para generarlo, o sube el archivo revisado."
        )

    mapeo = json.loads(ajustes.ARCHIVO_MAPEO.read_text(encoding="utf-8"))
    actividades = mapeo.get("actividades") or {}
    familias_ok = set(_opciones_familia())
    listas_ok = set(_opciones_lista(mapeo))

    wb = openpyxl.load_workbook(RUTA_EXCEL, data_only=True)
    if HOJA_MAPEO not in wb.sheetnames:
        raise SystemExit(f"ERROR: el archivo no tiene la hoja '{HOJA_MAPEO}'.")
    ws = wb[HOJA_MAPEO]

    cambios, errores, desconocidas = [], [], 0
    for r in range(2, ws.max_row + 1):
        clave = ws.cell(r, 1).value
        if not clave or not str(clave).strip():
            continue
        clave = str(clave).strip()
        if clave not in actividades:
            desconocidas += 1
            continue

        familia = str(ws.cell(r, 3).value or "").strip()
        lista = str(ws.cell(r, 4).value or "").strip()

        if familia not in familias_ok:
            errores.append(f"  fila {r}: familia {familia!r} no existe. "
                           f"Validas: {', '.join(sorted(familias_ok))}")
            continue
        if lista not in listas_ok:
            errores.append(f"  fila {r}: lista {lista!r} no existe en el tablero.")
            continue

        actual = actividades[clave]
        if actual.get("familia") != familia or actual.get("lista") != lista:
            cambios.append((actual.get("actividad", clave),
                            actual.get("familia"), familia,
                            actual.get("lista"), lista))
            actual["familia"] = familia
            actual["lista"] = lista

    if errores:
        print("ERROR: el cuadro tiene valores invalidos. NO se aplico nada:\n"
              + "\n".join(errores), file=sys.stderr)
        return 1

    if desconocidas:
        print(f"AVISO: {desconocidas} filas con una clave que ya no esta en el "
              f"cronograma; se ignoraron.")

    if not cambios:
        print("Revisado: no habia nada que corregir. El mapeo queda como estaba.")
        resumen = "sin cambios"
    else:
        print(f"Se aplican {len(cambios)} cambios:")
        for actividad, fam_v, fam_n, lst_v, lst_n in cambios:
            print(f"  · {actividad}")
            if fam_v != fam_n:
                print(f"      familia: {fam_v} -> {fam_n}")
            if lst_v != lst_n:
                print(f"      lista:   {lst_v} -> {lst_n}")
        resumen = f"{len(cambios)} actividades remapeadas"
        if not dry_run:
            ajustes.ARCHIVO_MAPEO.write_text(
                json.dumps(mapeo, ensure_ascii=False, indent=2) + "\n",
                encoding="utf-8")
            print(f"\nEscrito: {ajustes.ARCHIVO_MAPEO.name}")

    if dry_run:
        print("\n(DRY-RUN: no se escribio nada.)")

    salida = os.environ.get("GITHUB_OUTPUT")
    if salida:
        with open(salida, "a", encoding="utf-8") as f:
            f.write(f"resumen={resumen}\n")
    return 0


def main() -> int:
    ap = argparse.ArgumentParser(
        description="Cuadro de verificacion del mapeo, en Excel con desplegables.")
    grupo = ap.add_mutually_exclusive_group(required=True)
    grupo.add_argument("--generar", action="store_true",
                       help="Crea el cuadro a partir de mapeo.json.")
    grupo.add_argument("--aplicar", action="store_true",
                       help="Lee el cuadro revisado y actualiza mapeo.json.")
    ap.add_argument("--dry-run", action="store_true",
                    help="Con --aplicar: muestra los cambios sin escribir.")
    args = ap.parse_args()

    if args.generar:
        generar()
        return 0
    return aplicar(args.dry_run)


if __name__ == "__main__":
    sys.exit(main())
